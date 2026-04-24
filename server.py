#!/usr/bin/env python3
from flask import Flask, request, send_file, jsonify
import os, io, json
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__, static_folder='.', static_url_path='')

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, '오더지_양식.xlsx')
ITEMS_PATH    = os.path.join(BASE_DIR, 'items.json')
ADMIN_PW      = os.environ.get('ADMIN_PW', '1234')

# 기본 필드 컬럼 번호 (1-based, 최신 양식 검증 완료)
BASE = {
    'no':1, 'booth':2, 'company':3, 'payment':4, 'subtotal':6,
    'memo':244,   # IJ
    'bt':245,     # IK - 입금방법
    'bu':246,     # IL - 세금계산서
    'bv':247,     # IM - 담당자
}

ROW_PRE_START=13; ROW_PRE_END=63
ROW_POST_START=67; ROW_POST_END=71

def load_items():
    with open(ITEMS_PATH, encoding='utf-8') as f:
        return json.load(f)

def save_items(items):
    with open(ITEMS_PATH, 'w', encoding='utf-8') as f:
        json.dump(items, f, ensure_ascii=False, indent=2)

def check_pw():
    return request.headers.get('X-Admin-PW','') == ADMIN_PW

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/admin')
def admin():
    return app.send_static_file('admin.html')

@app.route('/ping')
def ping():
    return jsonify({'status':'ok','template':os.path.exists(TEMPLATE_PATH)})

@app.route('/api/items', methods=['GET'])
def get_items():
    return jsonify(load_items())

@app.route('/api/auth', methods=['POST'])
def auth():
    pw = request.get_json(force=True).get('pw','')
    return jsonify({'ok': pw == ADMIN_PW})

@app.route('/api/items', methods=['POST'])
def add_item():
    if not check_pw(): return jsonify({'error':'인증 실패'}), 401
    item = request.get_json()
    for f in ['key','col','cat','name']:
        if not str(item.get(f,'')).strip():
            return jsonify({'error': f'{f} 필드가 필요합니다'}), 400
    item.setdefault('detail',''); item.setdefault('code','')
    item['col'] = int(item['col'])
    items = load_items()
    if any(i['key'] == item['key'] for i in items):
        return jsonify({'error': f"키 {item['key']} 가 이미 존재합니다"}), 400
    items.append(item)
    items.sort(key=lambda x: x['col'])
    save_items(items)
    return jsonify({'ok':True,'items':items})

@app.route('/api/items/<key>', methods=['PUT'])
def update_item(key):
    if not check_pw(): return jsonify({'error':'인증 실패'}), 401
    data = request.get_json()
    items = load_items()
    idx = next((i for i,x in enumerate(items) if x['key']==key), None)
    if idx is None: return jsonify({'error':'품목 없음'}), 404
    items[idx].update({k:v for k,v in data.items() if k!='key'})
    if 'col' in data: items[idx]['col'] = int(data['col'])
    items.sort(key=lambda x: x['col'])
    save_items(items)
    return jsonify({'ok':True,'items':items})

@app.route('/api/items/<key>', methods=['DELETE'])
def delete_item(key):
    if not check_pw(): return jsonify({'error':'인증 실패'}), 401
    items = load_items()
    new = [i for i in items if i['key']!=key]
    if len(new)==len(items): return jsonify({'error':'품목 없음'}), 404
    save_items(new)
    return jsonify({'ok':True,'items':new})

@app.route('/export', methods=['POST'])
def export():
    try:
        data = request.get_json()
        orders = data.get('orders',[])
        if not orders: return jsonify({'error':'저장된 발주가 없습니다.'}), 400
        if not os.path.exists(TEMPLATE_PATH): return jsonify({'error':'양식 파일 없음'}), 400

        item_col_map = {i['key']:i['col'] for i in load_items()}

        with open(TEMPLATE_PATH,'rb') as f:
            buf = io.BytesIO(f.read())
        wb = load_workbook(buf)
        ws = wb.active

        period = (orders[0].get('period') or '').strip()
        if period: ws.cell(3,1).value = '행사기간 : ' + period

        pre  = [o for o in orders if o.get('type')!='post']
        post = [o for o in orders if o.get('type')=='post']

        def write(order, row):
            ws.cell(row, BASE['no']).value      = order.get('no') or ''
            ws.cell(row, BASE['booth']).value   = order.get('booth') or ''
            ws.cell(row, BASE['company']).value = order.get('company') or ''
            ws.cell(row, BASE['payment']).value = order.get('payment') or ''
            if order.get('subtotal'):
                ws.cell(row, BASE['subtotal']).value = int(order['subtotal'])
            for f in ('memo','bt','bu','bv'):
                if order.get(f): ws.cell(row, BASE[f]).value = order[f]
            for key, qty in (order.get('items') or {}).items():
                col = item_col_map.get(key)
                if col and qty and int(qty)>0:
                    ws.cell(row, col).value = int(qty)

        for i,o in enumerate(pre):
            r=ROW_PRE_START+i
            if r>ROW_PRE_END: break
            write(o,r)
        for i,o in enumerate(post):
            r=ROW_POST_START+i
            if r>ROW_POST_END: break
            write(o,r)

        out=io.BytesIO(); wb.save(out); out.seek(0)
        fname=f"오더지_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(out, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error':str(e)}), 500

if __name__=='__main__':
    port=int(os.environ.get('PORT',5000))
    print(f"서버: http://localhost:{port}")
    print(f"관리자: http://localhost:{port}/admin  (비번: {ADMIN_PW})")
    print(f"양식 파일: {'있음' if os.path.exists(TEMPLATE_PATH) else '없음 ⚠️'}")
    app.run(host='0.0.0.0', port=port, debug=False)
